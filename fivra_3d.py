# /////////////////////////////////////////////////////////////////////////////
# // Standard Gen5 VRTT Imports and Objects
# /////////////////////////////////////////////////////////////////////////////
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy
from math import log10
from Common.APIs import DataAPI
from Common.APIs import DisplayAPI
from Common.APIs import MainAPI
from Common.Models import TriggerModel
from Common.Models import FanModel
from Common.Models import RawDataModel
from Common.Models import DataModel
from Common.Models import DisplayModel
from Common.Enumerations import HorizontalScale
from Common.Enumerations import ScoplessChannel
from Common.Enumerations import Cursors
from Common.APIs import DDR5API
from Common.APIs import MeasurementItemsAPI
from Common.APIs import MeasurementAPI
from Common.APIs import GeneratorAPI
from Common.Enumerations import PowerState
from Common.Enumerations import Transition
from Common.Enumerations import Protocol
from Common.Enumerations import RailTestMode
from System import String, Char, Int32, UInt16, Boolean, Array, Byte, Double
from System.Collections.Generic import List
from System.Collections.Generic import Dictionary
from openpyxl.drawing.image import Image
from openpyxl.chart import ScatterChart, Reference, Series, SurfaceChart3D
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import utils
import clr
import os
import sys

dll_path = r"C:\GEN5\PythonScripts"
sys.path.append(dll_path)


# Intel Example Script  Version 1.5
#
#    "3D" Frequency and Duty Cycle Sweep Test for the Gen5 VR Test Tool
#
#    This script will sweep a range of frequencies on a given voltage rail
#    at a specified load step. Additionally, at each frequency step the script
#    will sweep the load step duty cycle from 10% to 90% in 10% increments.
#    Data for VMin/VMax, Vpeak-to-peak, and duty cycle
#    are collected at each frequency step, and are stored in an Excel file.
#    Scope shots of the worst points are taken and embedded. This script
#    uses the persistent voltage min/max with a dwell time of ~3 seconds.
#
#    Configurable parameters are listed below.
#
#    Output to Excel will be on a sheet titled "3D".
#
# Revision History
#
#    1.0     Initial version.
#    1.1     Removed dependency on MeasurementItems library. Added VMin/VMax
#            measurement to the output.
#    1.2     Added roll up and down options, formatting
#    1.3     Added drive rise/fall time options
#    1.4     Added a non loading sense rail to measure
#            Script will not try to load NOLOAD rails
#    1.5     Takes screen shots of worst cases, different formatting, uses persistent vmax/vmin to calculate
#    1.6     Fixed some math errors in the roll up/roll down algorithm.
#            Now uses a rise and fall slew rate for calculation.
#            If [fall_slew_rate] is equal to None then it is assumed to be equal to [rise_slew_rate]
#            If the current needs to be rolled up/down.
#               Measures VMax when current is rolled down and VMin when current is rolled up.
#    1.7     Added cooldowns, fixed vmax adjusted currents not being recorded.
#    1.8     Added tracking option and minor bug fixes


ScriptVersion = 1.8


def fivra_3d(target_voltage,
             start_frequency,
             end_frequency,
             sample_per_decade,
             nw_current_low,
             nw_current_high,
             ne_current_low,
             ne_current_high,
             sw_current_low,
             sw_current_high,
             se_current_low,
             se_current_high,
             nw_slew_rate,
             ne_slew_rate,
             sw_slew_rate,
             se_slew_rate,
             cooldown_delay,
             duty_cycle_list,
             debug=False,
             ):
    # ///////////////////////////////////////////////////////////////////////////////
    # // 3D Test Parameters - Edit these
    # ///////////////////////////////////////////////////////////////////////////////

    # [target_rail_1] is required, all other rails can be set to "NOLOAD" to ignore
    # Enables tracking that will auto adjust current level to match current setpoint
    enable_current_tracking = True

    target_rail_1 = "VCCFAEHVFIVRANW"  # Voltage rail to be tested
    target_rail_2 = "VCCFAEHVFIVRANE"  # Voltage rail to be tested
    target_rail_3 = "VCCFAEHVFIVRASW"  # Voltage rail to be tested
    target_rail_4 = "VCCFAEHVFIVRASE"  # Voltage rail to be tested
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_1 = "VCCFAEHVFIVRA"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_2 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_3 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_4 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_5 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_6 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_7 = "NOLOAD"
    # No load will be placed on this rail, but the voltage will be measured
    sense_rail_8 = "NOLOAD"
    


    sync_offset_2 = 0  # Phase offset for [target_rail_2] from [target_rail_1]
    sync_offset_3 = 0  # Phase offset for [target_rail_3] from [target_rail_1]
    sync_offset_4 = 0  # Phase offset for [target_rail_4] from [target_rail_1]

    target_voltage = target_voltage
    # Test voltage level for [target_rail_1](Volts)
    test_voltage_1 = target_voltage
    # Test voltage level for [target_rail_2](Volts)
    test_voltage_2 = target_voltage
    # Test voltage level for [target_rail_3](Volts)
    test_voltage_3 = target_voltage
    # Test voltage level for [target_rail_4](Volts)
    test_voltage_4 = target_voltage

    start_frequency = start_frequency  # Start Frequency(kHz)
    end_frequency = end_frequency  # End Frequency(kHz)
    # Number of sample steps (per decade)
    samples_per_decade = sample_per_decade

    nw_current_low = nw_current_low
    nw_current_high = nw_current_high
    ne_current_low = ne_current_low
    ne_current_high = ne_current_high
    sw_current_low = sw_current_low
    sw_current_high = sw_current_high
    se_current_low = se_current_low
    se_current_high = se_current_high

    test_current_low_1 = nw_current_low  # Load step IccMin(A)
    test_current_high_1 = nw_current_high  # Load step test_current_end(A)
    test_current_low_2 = ne_current_low  # Load step IccMin(A)
    test_current_high_2 = ne_current_high  # Load step test_current_end(A)
    test_current_low_3 = sw_current_low  # Load step IccMin(A)
    test_current_high_3 = sw_current_high  # Load step test_current_end(A)
    test_current_low_4 = se_current_low  # Load step IccMin(A)
    test_current_high_4 = se_current_high  # Load step test_current_end(A)

    nw_slew_rate = nw_slew_rate
    ne_slew_rate = ne_slew_rate
    sw_slew_rate = sw_slew_rate
    se_slew_rate = se_slew_rate
    rise_slew_rate_1 = nw_slew_rate  # Platform rising slew rate(A/us)
    # This will be assumed equal to [rise_slew_rate] if None (A/us)
    fall_slew_rate_1 = nw_slew_rate
    rise_slew_rate_2 = ne_slew_rate  # Platform rising slew rate(A/us)
    # This will be assumed equal to [rise_slew_rate] if None (A/us)
    fall_slew_rate_2 = ne_slew_rate
    rise_slew_rate_3 = sw_slew_rate  # Platform rising slew rate(A/us)
    # This will be assumed equal to [rise_slew_rate] if None (A/us)
    fall_slew_rate_3 = sw_slew_rate
    rise_slew_rate_4 = se_slew_rate  # Platform rising slew rate(A/us)
    # This will be assumed equal to [rise_slew_rate] if None (A/us)
    fall_slew_rate_4 = se_slew_rate

    nw_rise_time = (nw_current_high-nw_current_low)/nw_slew_rate*1000
    ne_rise_time = (ne_current_high-ne_current_low)/ne_slew_rate*1000
    sw_rise_time = (sw_current_high-sw_current_low)/sw_slew_rate*1000
    se_rise_time = (se_current_high-se_current_low)/se_slew_rate*1000

    drive_rise_time_1 = nw_rise_time  # Load step rise time(ns)
    drive_fall_time_1 = nw_rise_time  # Load step fall time(ns)
    drive_rise_time_2 = ne_rise_time  # Load step rise time(ns)
    drive_fall_time_2 = ne_rise_time  # Load step fall time(ns)
    drive_rise_time_3 = sw_rise_time  # Load step rise time(ns)
    drive_fall_time_3 = sw_rise_time  # Load step fall time(ns)
    drive_rise_time_4 = se_rise_time  # Load step rise time(ns)
    drive_fall_time_4 = se_rise_time  # Load step fall time(ns)

    # The test will turn off the load and pause after each DIMM measurement for [cooldown_delay] before starting the next measurement
    # This time is per each DIMM, duty_cycle, and frequency. This will greatly increase test time. A better cooling solution is preferred
    cooldown_delay = cooldown_delay

    # Set-up Duty Cycle list (units are %)
    duty_cycle_list = duty_cycle_list

    if debug == True:
        print(f'''[target_voltage]={target_voltage},
         [start_frequency]={start_frequency},
         [end_frequency]={end_frequency},
         [sample_per_decade]={sample_per_decade},
         [nw_current_low]={nw_current_low},
         [nw_current_high]={nw_current_high},
         [ne_current_low]={ne_current_low},
         [ne_current_high]={ne_current_high},
         [sw_current_low]={sw_current_low},
         [sw_current_high]={sw_current_high},
         [se_current_low]={se_current_low},
         [se_current_high]={se_current_high},
         [nw_slew_rate]={nw_slew_rate},
         [ne_slew_rate]={ne_slew_rate},
         [sw_slew_rate]={sw_slew_rate},
         [se_slew_rate]={se_slew_rate},
         [cooldown_delay]={cooldown_delay},
         [duty_cycle_list]={duty_cycle_list},
         [nw_rise_time]={nw_rise_time},
         [ne_rise_time]={ne_rise_time},
         [sw_rise_time]={sw_rise_time},
         [se_rise_time]={se_rise_time}'''
              )

    num_decades = int(log10(end_frequency) - log10(start_frequency))

    # Select a list of frequencies this method gives [samples_per_decade] linearly spaced per decade
    frequency_list = []
    for i in range(int(num_decades)):
        frequency_list.extend(numpy.linspace(
            10**(i), 10**(i+1), num=samples_per_decade, endpoint=True))

    # You can manually specify the frequencies in a list like this:
    # frequency_list = [1,2,3,4,5,6,7,8,9,10,20,30,40,50,60,70,80,90,100,200,300,400,500,600,700,800,900,1000]

    # For logarithmic spacing use this:
    # frequency_list = numpy.logspace(log10(start_frequency), log10(end_frequency), int(num_decades * samples_per_decade), True)

    output_file_name = "3D_4LC_Output" + \
        time.strftime("_%j_%M_%S", time.localtime()) + ".xlsx"

    # Power state parameters (target rail 1 only)
    enable_power_states = False
    ps_low = PowerState.Zero
    ps_high = PowerState.Zero
    # (us) Delay after the falling current edge before setting ps state(us)
    ps_low_delay = 3
    # (us) Delay after the falling current edge before setting ps state(us)
    ps_high_delay = 3

    # ///////////////////////////////////////////////////////////////////////////////
    # // END 3D Test Parameters
    # ///////////////////////////////////////////////////////////////////////////////

    generator_api = GeneratorAPI()
    data_api = DataAPI()
    display_api = DisplayAPI()
    measurement_api = MeasurementAPI()
    measurement_items_api = MeasurementItemsAPI()
    main_api = MainAPI()
    ddr = DDR5API()
    rails = generator_api.GetRails()
    minimum_up_time = 0
    ps_initialized = False
    
    # Switch back to SVDC mode and turn off the load...
    generator_api.Generator1SVSC(target_rail_1, 0, False)
    generator_api.Generator2SVSC(target_rail_2, 0, False)
    generator_api.Generator3SVSC(target_rail_3, 0, False)
    generator_api.Generator4SVSC(target_rail_4, 0, False)

    def adjust_horizontal_scale(frequency, min_pulses=3):
        period_us = (1 / (frequency * 1000)) * 1000000
        minimum_time = period_us * min_pulses
        minimum_time = minimum_time / 8  # minimum time per division (8)
        if minimum_time >= HorizontalScale.Scale5ms:
            display_api.SetHorizontalScale(HorizontalScale.Scale5ms)
            return HorizontalScale.Scale5ms
        elif minimum_time >= HorizontalScale.Scale2ms:
            display_api.SetHorizontalScale(HorizontalScale.Scale2ms)
            return HorizontalScale.Scale2ms
        elif minimum_time >= HorizontalScale.Scale1ms:
            display_api.SetHorizontalScale(HorizontalScale.Scale1ms)
            return HorizontalScale.Scale1ms
        elif minimum_time >= HorizontalScale.Scale500us:
            display_api.SetHorizontalScale(HorizontalScale.Scale500us)
            return HorizontalScale.Scale500us
        elif minimum_time >= HorizontalScale.Scale200us:
            display_api.SetHorizontalScale(HorizontalScale.Scale200us)
            return HorizontalScale.Scale200us
        elif minimum_time >= HorizontalScale.Scale80us:
            display_api.SetHorizontalScale(HorizontalScale.Scale80us)
            return HorizontalScale.Scale80us
        elif minimum_time >= HorizontalScale.Scale40us:
            display_api.SetHorizontalScale(HorizontalScale.Scale40us)
            return HorizontalScale.Scale40us
        elif minimum_time >= HorizontalScale.Scale20us:
            display_api.SetHorizontalScale(HorizontalScale.Scale20us)
            return HorizontalScale.Scale20us
        else:
            display_api.SetHorizontalScale(HorizontalScale.Scale10us)
            return HorizontalScale.Scale10us

    def needs_roll_up_down(low_current, high_current, frequency, duty_cycle, rise_slew_rate, fall_slew_rate, drive_rise_time, drive_fall_time, min_up_time=0):
        if fall_slew_rate is None:
            fall_slew_rate = rise_slew_rate

        rise_us = (abs(high_current - low_current) / (rise_slew_rate))
        fall_us = (abs(high_current - low_current) / (fall_slew_rate))
        vr_max_loadstep_min_time = min_up_time + rise_us + fall_us + min_up_time

        frequency_hz = frequency * 1000  # frequency in Hz
        period = 1 / frequency_hz  # s
        period_us = period * 1000000  # us
        rise_fall_time_available = ((period_us * (duty_cycle / 100)))

        # Not enough time for desired rise fall transistions
        return vr_max_loadstep_min_time >= rise_fall_time_available

    def adjust_current_levels(low_current, high_current, frequency, duty_cycle, rise_slew_rate, fall_slew_rate, drive_rise_time, drive_fall_time, roll_up=True, min_up_time=0):
        if fall_slew_rate is None:
            fall_slew_rate = rise_slew_rate

        if needs_roll_up_down(low_current, high_current, frequency, duty_cycle, rise_slew_rate, fall_slew_rate, drive_rise_time, drive_fall_time, min_up_time):
            frequency_hz = frequency * 1000  # frequency in Hz
            period = 1 / frequency_hz  # s
            period_us = period * 1000000  # us
            rise_fall_time_available = max(
                (period_us * (duty_cycle / 100)) - (min_up_time * 2), 0)
            current_part_1 = fall_slew_rate * \
                (rise_fall_time_available *
                 (rise_slew_rate / (rise_slew_rate + fall_slew_rate)))
            current_part_2 = rise_slew_rate * \
                (rise_fall_time_available *
                 (fall_slew_rate / (rise_slew_rate + fall_slew_rate)))
            current_step = (current_part_1 + current_part_2) / 2

            if roll_up:
                # Increase the lower current, but not beyond the higher current
                if low_current > high_current:
                    high_current = max(
                        low_current - current_step, high_current)
                else:
                    low_current = max(high_current - current_step, low_current)
            else:
                # Decrease the higher current, but not beyond the lower current
                if low_current > high_current:
                    low_current = min(high_current + current_step, low_current)
                else:
                    high_current = min(
                        low_current + current_step, high_current)
        return low_current, high_current

    rise_slew_rates = [rise_slew_rate_1, rise_slew_rate_2,
                       rise_slew_rate_3, rise_slew_rate_4]
    fall_slew_rates = [fall_slew_rate_1, fall_slew_rate_2,
                       fall_slew_rate_3, fall_slew_rate_4]
    load_rails = []
    rail_voltages = [test_voltage_1, test_voltage_2,
                     test_voltage_3, test_voltage_4]
    for idx, rail_name in enumerate([target_rail_1, target_rail_2, target_rail_3, target_rail_4]):
        if rail_name != "NOLOAD" and rail_name not in load_rails:
            load_rails.append(rail_name)
            generator_api.SetVoltageForRail(
                rail_name, rail_voltages[idx], Transition.Fast)

    test_rails = []
    for rail_name in [target_rail_1, target_rail_2, target_rail_3, target_rail_4, sense_rail_1, sense_rail_2, sense_rail_3, sense_rail_4, sense_rail_5, sense_rail_6, sense_rail_7, sense_rail_8]:
        if rail_name != "NOLOAD" and rail_name not in test_rails:
            test_rails.append(rail_name)

    # Set default trigger
    data_api.SetTrigger(4, 0, 0, 0, 0, False)

    # Create the output Excel file...
    wb = Workbook()
    worksheets = []
    raw_worksheets = []
    for rail_idx, test_rail in enumerate(test_rails):
        worksheets.append(wb.create_sheet(title="{}".format(test_rail)))
        raw_worksheets.append(wb.create_sheet(
            title="Raw 3D {}".format(test_rail)))

    # Set up Excel Doc:
    # Parameter disclosure...
    sheet_info = wb.create_sheet(title="3D Script Info")
    sheet_info.cell(column=1, row=1, value="Test Rail 1:")
    sheet_info.cell(column=2, row=1, value=target_rail_1)
    sheet_info.cell(column=1, row=2, value="Test Rail 2:")
    sheet_info.cell(column=2, row=2, value=target_rail_2)
    sheet_info.cell(column=1, row=3, value="Test Rail 3:")
    sheet_info.cell(column=2, row=3, value=target_rail_3)
    sheet_info.cell(column=1, row=4, value="Test Rail 4:")
    sheet_info.cell(column=2, row=4, value=target_rail_4)
    sheet_info.cell(column=1, row=5, value="Sense Rail 1")
    sheet_info.cell(column=2, row=5, value=sense_rail_1)
    sheet_info.cell(column=1, row=6, value="Voltage 1(V):")
    sheet_info.cell(column=2, row=6, value=test_voltage_1)
    sheet_info.cell(column=1, row=7, value="Voltage 2(V):")
    sheet_info.cell(column=2, row=7, value=test_voltage_2)
    sheet_info.cell(column=1, row=8, value="Voltage 3(V):")
    sheet_info.cell(column=2, row=8, value=test_voltage_3)
    sheet_info.cell(column=1, row=9, value="Voltage 4(V):")
    sheet_info.cell(column=2, row=9, value=test_voltage_4)
    sheet_info.cell(column=1, row=10, value="Start Current 1(A):")
    sheet_info.cell(column=2, row=10, value=test_current_low_1)
    sheet_info.cell(column=1, row=11, value="End Current 1(A):")
    sheet_info.cell(column=2, row=11, value=test_current_high_1)
    sheet_info.cell(column=1, row=12, value="Start Current 2(A):")
    sheet_info.cell(column=2, row=12, value=test_current_low_2)
    sheet_info.cell(column=1, row=13, value="End Current 2(A):")
    sheet_info.cell(column=2, row=13, value=test_current_high_2)
    sheet_info.cell(column=1, row=14, value="Start Current 3(A):")
    sheet_info.cell(column=2, row=14, value=test_current_low_3)
    sheet_info.cell(column=1, row=15, value="End Current 3(A):")
    sheet_info.cell(column=2, row=15, value=test_current_high_3)
    sheet_info.cell(column=1, row=16, value="Start Current 4(A):")
    sheet_info.cell(column=2, row=16, value=test_current_low_4)
    sheet_info.cell(column=1, row=17, value="End Current 4(A):")
    sheet_info.cell(column=2, row=17, value=test_current_high_4)
    sheet_info.cell(column=1, row=18, value="Drive Rise Time 1(ns):")
    sheet_info.cell(column=2, row=18, value=drive_rise_time_1)
    sheet_info.cell(column=1, row=19, value="Drive Fall Time 1(ns):")
    sheet_info.cell(column=2, row=19, value=drive_fall_time_1)
    sheet_info.cell(column=1, row=20, value="Drive Rise Time 2(ns):")
    sheet_info.cell(column=2, row=20, value=drive_rise_time_2)
    sheet_info.cell(column=1, row=21, value="Drive Fall Time 2(ns):")
    sheet_info.cell(column=2, row=21, value=drive_fall_time_2)
    sheet_info.cell(column=1, row=22, value="Drive Rise Time 3(ns):")
    sheet_info.cell(column=2, row=22, value=drive_rise_time_3)
    sheet_info.cell(column=1, row=23, value="Drive Fall Time 3(ns):")
    sheet_info.cell(column=2, row=23, value=drive_fall_time_3)
    sheet_info.cell(column=1, row=24, value="Drive Rise Time 4(ns):")
    sheet_info.cell(column=2, row=24, value=drive_rise_time_4)
    sheet_info.cell(column=1, row=25, value="Drive Fall Time 4(ns):")
    sheet_info.cell(column=2, row=25, value=drive_fall_time_4)
    sheet_info.cell(column=1, row=26,
                    value="Max Rail 1 Rising Slew Rate(A/us)")
    sheet_info.cell(column=2, row=26, value=rise_slew_rate_1)
    sheet_info.cell(column=1, row=27,
                    value="Max Rail 1 Falling Slew Rate(A/us)")
    sheet_info.cell(column=2, row=27, value=fall_slew_rate_1)
    sheet_info.cell(column=1, row=28,
                    value="Max Rail 2 Rising Slew Rate(A/us)")
    sheet_info.cell(column=2, row=28, value=rise_slew_rate_2)
    sheet_info.cell(column=1, row=29,
                    value="Max Rail 2 Falling Slew Rate(A/us)")
    sheet_info.cell(column=2, row=29, value=fall_slew_rate_2)
    sheet_info.cell(column=1, row=30,
                    value="Max Rail 3 Rising Slew Rate(A/us)")
    sheet_info.cell(column=2, row=30, value=rise_slew_rate_3)
    sheet_info.cell(column=1, row=31,
                    value="Max Rail 3 Falling Slew Rate(A/us)")
    sheet_info.cell(column=2, row=31, value=fall_slew_rate_3)
    sheet_info.cell(column=1, row=32,
                    value="Max Rail 4 Rising Slew Rate(A/us)")
    sheet_info.cell(column=2, row=32, value=rise_slew_rate_4)
    sheet_info.cell(column=1, row=33,
                    value="Max Rail 4 Falling Slew Rate(A/us)")
    sheet_info.cell(column=2, row=33, value=fall_slew_rate_4)

    c = sheet_info.cell(column=1, row=34, value="Start Freqency (kHz):")
    sheet_info.column_dimensions[c.column_letter].width = len(str(c.value))
    sheet_info.cell(column=2, row=34, value=start_frequency)
    sheet_info.cell(column=1, row=35, value="End Frequency (kHz):")
    sheet_info.cell(column=2, row=35, value=end_frequency)
    sheet_info.cell(column=1, row=36, value="Sample per decade:")
    sheet_info.cell(column=2, row=36, value=samples_per_decade)
    sheet_info.cell(column=1, row=37, value="3D Script Version:")
    sheet_info.cell(column=2, row=37, value=ScriptVersion)

    sheet_info.cell(column=1, row=38, value="Sense Rail 2")
    sheet_info.cell(column=2, row=38, value=sense_rail_2)
    sheet_info.cell(column=1, row=39, value="Sense Rail 3")
    sheet_info.cell(column=2, row=39, value=sense_rail_3)
    sheet_info.cell(column=1, row=40, value="Sense Rail 4")
    sheet_info.cell(column=2, row=40, value=sense_rail_4)
    sheet_info.cell(column=1, row=41, value="Sense Rail 5")
    sheet_info.cell(column=2, row=41, value=sense_rail_5)
    sheet_info.cell(column=1, row=42, value="Sense Rail 6")
    sheet_info.cell(column=2, row=42, value=sense_rail_6)
    sheet_info.cell(column=1, row=43, value="Sense Rail 7")
    sheet_info.cell(column=2, row=43, value=sense_rail_7)
    sheet_info.cell(column=1, row=44, value="Sense Rail 8")
    sheet_info.cell(column=2, row=44, value=sense_rail_8)

    del wb["Sheet"]
    wb.save(filename=output_file_name)

    # Data column headers...
    for ws in raw_worksheets:
        ws.cell(column=5, row=1, value="Set Frequency(kHz)")
        ws.cell(column=6, row=1, value="Frequency(kHz)")
        ws.cell(column=7, row=1, value="DutyCycle(%)")

        ws.cell(column=9, row=1, value="Min Voltage(V)")
        ws.cell(column=10, row=1, value="Set High Current(A)")
        ws.cell(column=11, row=1, value="Set Low Current(A)")
        ws.cell(column=12, row=1, value="Measured High Current Level(A)")
        ws.cell(column=13, row=1, value="Measured Low Current Level(A)")

        ws.cell(column=15, row=1, value="Max Voltage(V)")
        ws.cell(column=16, row=1, value="Set High Current(A)")
        ws.cell(column=17, row=1, value="Set Low Current(A)")
        ws.cell(column=18, row=1, value="Measured High Current Level(A)")
        ws.cell(column=19, row=1, value="Measured Low Current Level(A)")

        ws.cell(column=1, row=1, value="Worst VMin(V)")
        c = ws.cell(column=1, row=2, value="Worst VMin Frequency(kHz)")
        ws.column_dimensions[c.column_letter].width = len(str(c.value))
        ws.cell(column=1, row=3, value="Worst VMin Duty Cycle(%)")
        ws.cell(column=1, row=4, value="Worst VMin High Current(A)")
        ws.cell(column=1, row=5, value="Worst VMin Low Current(A)")

        ws.cell(column=1, row=7, value="Worst VMax(V)")
        ws.cell(column=1, row=8, value="Worst VMax Frequency(kHz)")
        ws.cell(column=1, row=9, value="Worst VMax Duty Cycle(%)")
        ws.cell(column=1, row=10, value="Worst VMax High Current(A)")
        ws.cell(column=1, row=11, value="Worst VMax Low Current(A)")

    vmin_graph_col = "A"
    vmax_graph_col = "A"
    for ws in worksheets:
        for duty_cycle_idx, duty_cycle in enumerate(duty_cycle_list):
            c = ws.cell(column=5 + duty_cycle_idx, row=4, value=duty_cycle)
            ws.cell(column=5 + duty_cycle_idx, row=5 + len(frequency_list) + 3,
                    value="=MIN({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))
            ws.cell(column=5 + duty_cycle_idx, row=5 + len(frequency_list) + 4,
                    value="=MAX({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))

            c = ws.cell(column=5 + len(duty_cycle_list) + 6 +
                        duty_cycle_idx, row=4, value=duty_cycle)
            ws.cell(column=5 + len(duty_cycle_list) + 6 + duty_cycle_idx, row=5 + len(frequency_list) + 3,
                    value="=MIN({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))
            ws.cell(column=5 + len(duty_cycle_list) + 6 + duty_cycle_idx, row=6 + len(frequency_list) + 3,
                    value="=MAX({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))

            c = ws.cell(column=5 + 2*(len(duty_cycle_list) + 6) +
                        duty_cycle_idx, row=4, value=duty_cycle)
            ws.cell(column=5 + 2*(len(duty_cycle_list) + 6) + duty_cycle_idx, row=5 + len(frequency_list) + 3,
                    value="=MIN({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))
            ws.cell(column=5 + 2*(len(duty_cycle_list) + 6) + duty_cycle_idx, row=6 + len(frequency_list) + 3,
                    value="=MAX({}{}:{}{})".format(c.column_letter, 5, c.column_letter, 5 + len(frequency_list)))
        # Summary table headers, VMIN, VMAX, PkPk

        ws.cell(column=5, row=3, value="Duty Cycle(%)")
        ws.cell(column=5 + len(duty_cycle_list) +
                6, row=3, value="Duty Cycle(%)")
        ws.cell(column=5 + 2 * (len(duty_cycle_list) + 6),
                row=3, value="Duty Cycle(%)")

        ws.cell(column=5, row=2, value="VMIN Values")
        ws.cell(column=5 + len(duty_cycle_list) +
                6, row=2, value="VMAX Values")
        ws.cell(column=5 + 2*(len(duty_cycle_list) + 6),
                row=2, value="Peak to Peaks")

        for frequency_idx, frequency in enumerate(frequency_list):
            ws.cell(column=4, row=5 + frequency_idx, value=frequency)
            ws.cell(column=4 + len(duty_cycle_list) + 6,
                    row=5 + frequency_idx, value=frequency)
            ws.cell(column=4 + 2*(len(duty_cycle_list) + 6),
                    row=5 + frequency_idx, value=frequency)

        ws.cell(column=3, row=5, value="Set Frequency(kHz)")
        c = ws.cell(column=3 + len(duty_cycle_list) + 6,
                    row=5, value="Set Frequency(kHz)")
        ws.cell(column=3 + 2*(len(duty_cycle_list) + 6),
                row=5, value="Set Frequency(kHz)")
        ws.column_dimensions[c.column_letter].width = len(str(c.value))

        ws.cell(column=4, row=5 + len(frequency_list) + 3, value="MIN")
        c = ws.cell(column=4, row=5 + len(frequency_list) + 4, value="MAX")
        vmin_graph_col = c.column_letter  # Column letter to start 3d graph

        ws.cell(column=4 + len(duty_cycle_list) + 6,
                row=5 + len(frequency_list) + 3, value="MIN")
        c = ws.cell(column=4 + len(duty_cycle_list) + 6,
                    row=5 + len(frequency_list) + 4, value="MAX")
        vmax_graph_col = c.column_letter  # Column letter to start 3d graph

        ws.cell(column=4 + 2*(len(duty_cycle_list) + 6),
                row=5 + len(frequency_list) + 3, value="MIN")
        ws.cell(column=4 + 2*(len(duty_cycle_list) + 6),
                row=5 + len(frequency_list) + 4, value="MAX")
        # Not graphed so dont need column letter

    # Set the first row for data_api output...
    RowIndex = 2

    # Set Fan Speed to Max... enable if you are having thermal problems, this gets loud
    # data_api.SetFanSpeed(99)

    sample_number = 0  # For calculating percent complete

    worst_vmin_frequency = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmin_duty_cycle = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmax_frequency = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmax_duty_cycle = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmins = [100, 100, 100, 100, 100, 100, 100, 100, 100, 100, 100, 100]
    worst_pkpk = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmaxes = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmax_high_currents = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmax_low_currents = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmin_high_currents = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    worst_vmin_low_currents = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    for frequency_idx, frequency in enumerate(frequency_list):
        # Set the horizontal scale such that at least 2 complete
        # cycles are displayed at the current frequency...
        horizontal_scale = adjust_horizontal_scale(frequency)

        for duty_cycle_idx, duty_cycle in enumerate(duty_cycle_list):

            sample_number = sample_number + 1  # For percent complete only
            max_step_limited = False
            max_step_limited = max_step_limited or needs_roll_up_down(test_current_low_1, test_current_high_1, frequency, duty_cycle,
                                                                      rise_slew_rate_1, fall_slew_rate_1, drive_rise_time_1, drive_fall_time_1)
            if target_rail_2 != "NOLOAD":
                max_step_limited = max_step_limited or needs_roll_up_down(test_current_low_2, test_current_high_2, frequency, duty_cycle,
                                                                          rise_slew_rate_2, fall_slew_rate_2, drive_rise_time_2, drive_fall_time_2)
            if target_rail_3 != "NOLOAD":
                max_step_limited = max_step_limited or needs_roll_up_down(test_current_low_3, test_current_high_3, frequency, duty_cycle,
                                                                          rise_slew_rate_3, fall_slew_rate_3, drive_rise_time_3, drive_fall_time_3)
            if target_rail_4 != "NOLOAD":
                max_step_limited = max_step_limited or needs_roll_up_down(test_current_low_4, test_current_high_4, frequency, duty_cycle,
                                                                          rise_slew_rate_4, fall_slew_rate_4, drive_rise_time_4, drive_fall_time_4)
            adjusted_low_currents = [0] * len(test_rails)
            adjusted_high_currents = [0] * len(test_rails)

            for rail_idx, rail_name in enumerate(test_rails):
                if rail_name == "NOLOAD":
                    continue

                # Roll up for VMIN capture
                # Does nothing if slew rates can be met at this duty cycle and frequency

                adjusted_low_1, adjusted_high_1 = adjust_current_levels(test_current_low_1, test_current_high_1, frequency, duty_cycle,
                                                                        rise_slew_rate_1, fall_slew_rate_1, drive_rise_time_1, drive_fall_time_1, roll_up=True, min_up_time=minimum_up_time)
                adjusted_high_currents[rail_idx] = adjusted_high_1
                adjusted_low_currents[rail_idx] = adjusted_low_1
                if target_rail_2 != "NOLOAD":
                    adjusted_low_2, adjusted_high_2 = adjust_current_levels(test_current_low_2, test_current_high_2, frequency, duty_cycle,
                                                                            rise_slew_rate_2, fall_slew_rate_2, drive_rise_time_2, drive_fall_time_2, roll_up=True, min_up_time=minimum_up_time)
                    adjusted_high_currents[rail_idx] = adjusted_high_2
                    adjusted_low_currents[rail_idx] = adjusted_low_2
                if target_rail_3 != "NOLOAD":
                    adjusted_low_3, adjusted_high_3 = adjust_current_levels(test_current_low_3, test_current_high_3, frequency, duty_cycle,
                                                                            rise_slew_rate_3, fall_slew_rate_3, drive_rise_time_3, drive_fall_time_3, roll_up=True, min_up_time=minimum_up_time)
                    adjusted_high_currents[rail_idx] = adjusted_high_3
                    adjusted_low_currents[rail_idx] = adjusted_low_3
                if target_rail_4 != "NOLOAD":
                    adjusted_low_4, adjusted_high_4 = adjust_current_levels(test_current_low_4, test_current_high_4, frequency, duty_cycle,
                                                                            rise_slew_rate_4, fall_slew_rate_4, drive_rise_time_4, drive_fall_time_4, roll_up=True, min_up_time=minimum_up_time)
                    adjusted_high_currents[rail_idx] = adjusted_high_4
                    adjusted_low_currents[rail_idx] = adjusted_low_4

                if rail_name not in load_rails:
                    adjusted_high_currents[rail_idx] = 0
                    adjusted_low_currents[rail_idx] = 0

                # Display it on the scope view...
                if (rail_name != target_rail_1):
                    display_api.SetChannel(ScoplessChannel.Ch3, True)
                    display_api.SetChannel(ScoplessChannel.Ch4, True)
                    display_api.Ch3_4Rail(target_rail_1, True)
                    display_api.SetChannel(ScoplessChannel.Ch1, True)
                    display_api.SetChannel(ScoplessChannel.Ch2, True)
                    display_api.Ch1_2Rail(rail_name, True)
                    generator_api.SetTrackingEnabled(
                        rail_name, enable_current_tracking)
                    generator_api.SetTrackingEnabled(
                        target_rail_1, enable_current_tracking)

                else:
                    display_api.SetChannel(ScoplessChannel.Ch3, True)
                    display_api.SetChannel(ScoplessChannel.Ch4, True)
                    display_api.Ch3_4Rail(target_rail_1, True)
                    generator_api.SetTrackingEnabled(
                        target_rail_1, enable_current_tracking)

                time.sleep(2)
                # Set the load current step size, frequency, ramp time, and duty cycle.
                generator_api.Generator1DualSlopeSVDC(
                    target_rail_1, adjusted_high_1, adjusted_low_1, frequency, duty_cycle, drive_rise_time_1, drive_fall_time_1, True)
                if (enable_power_states and ps_initialized == False):
                    generator_api.SetPS_SVDC(
                        target_rail_1, enable_power_states, ps_low, ps_high, ps_low_delay, ps_high_delay)
                    ps_initialized = True
                if target_rail_2 != "NOLOAD":
                    generator_api.Generator2DualSlopeSVDC_SynchON(target_rail_2, adjusted_high_2, adjusted_low_2, frequency, duty_cycle,
                                                                  drive_rise_time_2, drive_fall_time_2, True, sync_offset_2)
                if target_rail_3 != "NOLOAD":
                    generator_api.Generator3DualSlopeSVDC_SynchON(target_rail_3, adjusted_high_3, adjusted_low_3, frequency, duty_cycle,
                                                                  drive_rise_time_3, drive_fall_time_3, True, sync_offset_3)
                if target_rail_4 != "NOLOAD":
                    generator_api.Generator4DualSlopeSVDC_SynchON(target_rail_4, adjusted_high_4, adjusted_low_4, frequency, duty_cycle,
                                                                  drive_rise_time_4, drive_fall_time_4, True, sync_offset_4)

                # Set up all required measurements...
                measurement_api.ClearAllMeasurements()

                measurement_api.MeasureCurrentFrequency(rail_name)
                measurement_api.MeasurePersistentVoltageMinMax(rail_name)
                measurement_api.MeasureVoltageFrequency(rail_name)
                measurement_api.MeasureCurrentStateLevel(rail_name)
                time.sleep(1)
                measurement_api.ResetPersistentVoltageMeasurement(rail_name)
                # Wait 3 seconds for measurements to catch up...
                time.sleep(2)

                # Grab the measurements from the GUI
                if rail_name not in load_rails:
                    Frequency = 0
                else:
                    a = measurement_items_api.GetCurrentFrequencyOnce(
                        rail_name, -1)
                    b = a[0]
                    c = b.Value
                    Frequency = c
                # Grab the measurements from the GUI
                # Measure frequency

                # Measure persistent voltage min max
                vmin_vmax_measurement = measurement_items_api.GetPersistentVoltageMinMaxOnce(
                    rail_name, -1)
                vmin_vmax_measurement = measurement_items_api.GetPersistentVoltageMinMaxOnce(
                    rail_name, -1)
                vmax = vmin_vmax_measurement[0].Value
                vmin = vmin_vmax_measurement[1].Value
                # Measure current levels
                current_state_measurement = measurement_items_api.GetCurrentStateLevelsOnce(
                    rail_name, -1)
                current_low_level = current_state_measurement[1].Value
                current_high_level = current_state_measurement[0].Value

                raw_worksheets[rail_idx].cell(
                    column=9, row=RowIndex, value=vmin)
                raw_worksheets[rail_idx].cell(
                    column=10, row=RowIndex, value=adjusted_high_currents[rail_idx])
                raw_worksheets[rail_idx].cell(
                    column=11, row=RowIndex, value=adjusted_low_currents[rail_idx])
                raw_worksheets[rail_idx].cell(
                    column=12, row=RowIndex, value=float(current_high_level))
                raw_worksheets[rail_idx].cell(
                    column=13, row=RowIndex, value=float(current_low_level))

                # New record for worst vdroop/undershoot
                if vmin < worst_vmins[rail_idx]:
                    worst_vmins[rail_idx] = vmin
                    worst_vmin_frequency[rail_idx] = frequency
                    worst_vmin_duty_cycle[rail_idx] = duty_cycle
                    worst_vmin_high_currents[rail_idx] = adjusted_high_currents[rail_idx]
                    worst_vmin_low_currents[rail_idx] = adjusted_low_currents[rail_idx]

                if max_step_limited:
                    # Replaces the VMAX taken while rolled down
                    # Roll down for VMAX capture

                    adjusted_low_1, adjusted_high_1 = adjust_current_levels(test_current_low_1, test_current_high_1, frequency, duty_cycle,
                                                                            rise_slew_rate_1, fall_slew_rate_1, drive_rise_time_1, drive_fall_time_1, roll_up=False, min_up_time=minimum_up_time)
                    adjusted_high_currents[rail_idx] = adjusted_high_1
                    adjusted_low_currents[rail_idx] = adjusted_low_1
                    if target_rail_2 != "NOLOAD":
                        adjusted_low_2, adjusted_high_2 = adjust_current_levels(test_current_low_2, test_current_high_2, frequency, duty_cycle,
                                                                                rise_slew_rate_2, fall_slew_rate_2, drive_rise_time_2, drive_fall_time_2, roll_up=False, min_up_time=minimum_up_time)
                        adjusted_high_currents[rail_idx] = adjusted_high_2
                        adjusted_low_currents[rail_idx] = adjusted_low_2
                    if target_rail_3 != "NOLOAD":
                        adjusted_low_3, adjusted_high_3 = adjust_current_levels(test_current_low_3, test_current_high_3, frequency, duty_cycle,
                                                                                rise_slew_rate_3, fall_slew_rate_3, drive_rise_time_3, drive_fall_time_3, roll_up=False, min_up_time=minimum_up_time)
                        adjusted_high_currents[rail_idx] = adjusted_high_3
                        adjusted_low_currents[rail_idx] = adjusted_low_3
                    if target_rail_4 != "NOLOAD":
                        adjusted_low_4, adjusted_high_4 = adjust_current_levels(test_current_low_4, test_current_high_4, frequency, duty_cycle,
                                                                                rise_slew_rate_4, fall_slew_rate_4, drive_rise_time_4, drive_fall_time_4, roll_up=False, min_up_time=minimum_up_time)
                        adjusted_high_currents[rail_idx] = adjusted_high_4
                        adjusted_low_currents[rail_idx] = adjusted_low_4

                    if rail_name not in load_rails:
                        adjusted_high_currents[rail_idx] = 0
                        adjusted_low_currents[rail_idx] = 0
                    # Set the load current step size, frequency, ramp time, and duty cycle.
                    generator_api.Generator1DualSlopeSVDC(
                        target_rail_1, adjusted_high_1, adjusted_low_1, frequency, duty_cycle, drive_rise_time_1, drive_fall_time_1, True)
                    if target_rail_2 != "NOLOAD":
                        generator_api.Generator2DualSlopeSVDC_SynchON(target_rail_2, adjusted_high_2, adjusted_low_2, frequency, duty_cycle,
                                                                      drive_rise_time_2, drive_fall_time_2, True, sync_offset_2)
                    if target_rail_3 != "NOLOAD":
                        generator_api.Generator3DualSlopeSVDC_SynchON(target_rail_3, adjusted_high_3, adjusted_low_3, frequency, duty_cycle,
                                                                      drive_rise_time_3, drive_fall_time_3, True, sync_offset_3)
                    if target_rail_4 != "NOLOAD":
                        generator_api.Generator4DualSlopeSVDC_SynchON(target_rail_4, adjusted_high_4, adjusted_low_4, frequency, duty_cycle,
                                                                      drive_rise_time_4, drive_fall_time_4, True, sync_offset_4)

                    # Set up all required measurements...
                    measurement_api.ResetPersistentVoltageMeasurement(
                        rail_name)
                    time.sleep(2)
                    # Grab the measurements from the GUI
                    # Measure frequency
                    Frequency = measurement_items_api.GetCurrentFrequencyOnce(rail_name)[
                        0].Value
                    # Measure persistent voltage min max
                    vmin_vmax_measurement = measurement_items_api.GetPersistentVoltageMinMaxOnce(
                        rail_name)
                    vmax = vmin_vmax_measurement[0].Value
                    vmin = vmin_vmax_measurement[1].Value
                    # Measure current levels
                    current_state_measurement = measurement_items_api.GetCurrentStateLevelsOnce(
                        rail_name)
                    current_low_level = current_state_measurement[1].Value
                    current_high_level = current_state_measurement[0].Value

                # New record for worst overshoot
                if vmax > worst_vmaxes[rail_idx]:
                    worst_vmaxes[rail_idx] = vmax
                    worst_vmax_frequency[rail_idx] = frequency
                    worst_vmax_duty_cycle[rail_idx] = duty_cycle
                    worst_vmax_high_currents[rail_idx] = adjusted_high_currents[rail_idx]
                    worst_vmax_low_currents[rail_idx] = adjusted_low_currents[rail_idx]

                if worst_vmaxes[rail_idx] - worst_vmins[rail_idx] > worst_pkpk[rail_idx]:
                    worst_pkpk[rail_idx] = worst_vmaxes[rail_idx] - \
                        worst_vmins[rail_idx]
                # Raw Data
                raw_worksheets[rail_idx].cell(
                    column=5, row=RowIndex, value=frequency)
                raw_worksheets[rail_idx].cell(
                    column=6, row=RowIndex, value=Frequency)
                raw_worksheets[rail_idx].cell(
                    column=7, row=RowIndex, value=duty_cycle)

                raw_worksheets[rail_idx].cell(
                    column=15, row=RowIndex, value=vmax)
                raw_worksheets[rail_idx].cell(
                    column=16, row=RowIndex, value=adjusted_high_currents[rail_idx])
                raw_worksheets[rail_idx].cell(
                    column=17, row=RowIndex, value=adjusted_low_currents[rail_idx])
                raw_worksheets[rail_idx].cell(
                    column=18, row=RowIndex, value=current_high_level)
                raw_worksheets[rail_idx].cell(
                    column=19, row=RowIndex, value=current_low_level)

                worksheets[rail_idx].cell(
                    column=5 + duty_cycle_idx, row=5 + frequency_idx, value=vmin)
                worksheets[rail_idx].cell(
                    column=5 + len(duty_cycle_list) + 6 + duty_cycle_idx, row=5 + frequency_idx, value=vmax)
                worksheets[rail_idx].cell(
                    column=5 + (2 * (len(duty_cycle_list) + 6)) + duty_cycle_idx, row=5 + frequency_idx, value=vmax-vmin)

            if cooldown_delay > 0:
                print("Cooling down for {} seconds...".format(cooldown_delay))
                generator_api.Generator1DualSlopeSVDC(
                    target_rail_1, 0, 0, frequency, duty_cycle, drive_rise_time_1, drive_fall_time_1, False)
                if target_rail_2 != "NOLOAD":
                    generator_api.Generator2DualSlopeSVDC_SynchON(target_rail_2, 0, 0, frequency, duty_cycle,
                                                                  drive_rise_time_2, drive_fall_time_2, False, sync_offset_2)
                if target_rail_3 != "NOLOAD":
                    generator_api.Generator3DualSlopeSVDC_SynchON(target_rail_3, 0, 0, frequency, duty_cycle,
                                                                  drive_rise_time_3, drive_fall_time_3, False, sync_offset_3)
                if target_rail_4 != "NOLOAD":
                    generator_api.Generator4DualSlopeSVDC_SynchON(target_rail_4, 0, 0, frequency, duty_cycle,
                                                                  drive_rise_time_4, drive_fall_time_4, False, sync_offset_4)
                time.sleep(cooldown_delay)

            # Advance to the next data row...
            RowIndex = RowIndex + 1
            # Save the output file...
            wb.save(filename=output_file_name)

    for rail_idx, rail_name in enumerate(test_rails):
        raw_worksheets[rail_idx].cell(
            column=2, row=1, value=worst_vmins[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=2, value=worst_vmin_frequency[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=3, value=worst_vmin_duty_cycle[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=4, value=worst_vmin_high_currents[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=5, value=worst_vmin_low_currents[rail_idx])

        raw_worksheets[rail_idx].cell(
            column=2, row=7, value=worst_vmaxes[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=8, value=worst_vmax_frequency[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=9, value=worst_vmax_duty_cycle[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=10, value=worst_vmax_high_currents[rail_idx])
        raw_worksheets[rail_idx].cell(
            column=2, row=11, value=worst_vmax_low_currents[rail_idx])

        raw_worksheets[rail_idx].cell(
            column=2, row=13, value=worst_pkpk[rail_idx])

        # Block Data
        worksheets[rail_idx].cell(column=1, row=4, value="Worst VMin(V)")
        c = worksheets[rail_idx].cell(
            column=1, row=5, value="Worst VMin Frequency(kHz)")
        worksheets[rail_idx].column_dimensions[c.column_letter].width = len(
            str(c.value))
        worksheets[rail_idx].cell(
            column=1, row=6, value="Worst VMin Duty Cycle(%)")
        c = worksheets[rail_idx].cell(
            column=1, row=7, value="Worst VMin High Current(A)")
        worksheets[rail_idx].column_dimensions[c.column_letter].width = len(
            str(c.value))
        worksheets[rail_idx].cell(
            column=1, row=8, value="Worst VMin Low Current(A)")

        worksheets[rail_idx].cell(column=2, row=4, value=worst_vmins[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=5, value=worst_vmin_frequency[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=6, value=worst_vmin_duty_cycle[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=7, value=worst_vmin_high_currents[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=8, value=worst_vmin_low_currents[rail_idx])

        worksheets[rail_idx].cell(column=1, row=10, value="Worst VMax(V)")
        worksheets[rail_idx].cell(
            column=1, row=11, value="Worst VMax Frequency(kHz)")
        worksheets[rail_idx].cell(
            column=1, row=12, value="Worst VMax Duty Cycle(%)")
        worksheets[rail_idx].cell(
            column=1, row=13, value="Worst VMax High Current(A)")
        worksheets[rail_idx].cell(
            column=1, row=14, value="Worst VMax Low Current(A)")

        worksheets[rail_idx].cell(
            column=2, row=10, value=worst_vmaxes[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=11, value=worst_vmax_frequency[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=12, value=worst_vmax_duty_cycle[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=13, value=worst_vmax_high_currents[rail_idx])
        worksheets[rail_idx].cell(
            column=2, row=14, value=worst_vmax_low_currents[rail_idx])

        worksheets[rail_idx].cell(
            column=1, row=16, value="Worst Peak to Peak Voltage")
        worksheets[rail_idx].cell(column=2, row=16, value=worst_pkpk[rail_idx])

    for ws_idx, ws in enumerate(worksheets):
        # Add the graph of the voltage traces

        ws.conditional_formatting.add(
            "{}{}:{}{}".format(
                utils.cell.get_column_letter(5), 5, utils.cell.get_column_letter(5 + len(duty_cycle_list)), 5 + len(frequency_list)), ColorScaleRule(
                start_type="percentile", start_value=0, start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="percentile", end_value=100, end_color="FA696A"))

        ws.conditional_formatting.add(
            "{}{}:{}{}".format(
                utils.cell.get_column_letter(5 + len(duty_cycle_list) + 6), 5, utils.cell.get_column_letter(5 + len(duty_cycle_list) + 6 + len(duty_cycle_list)), 5 + len(frequency_list)), ColorScaleRule(
                start_type="percentile", start_value=0, start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="percentile", end_value=100, end_color="FA696A"))

        ws.conditional_formatting.add(
            "{}{}:{}{}".format(
                utils.cell.get_column_letter(5 + 2*(len(duty_cycle_list) + 6)), 5, utils.cell.get_column_letter(5 + len(duty_cycle_list) + 2*(len(duty_cycle_list) + 6)), 5 + len(frequency_list)), ColorScaleRule(
                start_type="percentile", start_value=0, start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="percentile", end_value=100, end_color="FA696A"))

        # 3D Surface
        vmin_surface = SurfaceChart3D()
        ref = Reference(
            ws, min_col=5, max_col=4 + len(duty_cycle_list), min_row=4, max_row=4 + len(frequency_list))
        labels = Reference(ws, min_col=4, min_row=5,
                           max_row=4 + len(frequency_list))
        vmin_surface.add_data(ref, titles_from_data=True)
        vmin_surface.set_categories(labels)
        vmin_surface.title = "3D VMIN"
        ws.add_chart(vmin_surface, "{}{}".format(
            vmin_graph_col, 5 + len(frequency_list) + 6))

        vmax_surface = SurfaceChart3D()
        ref = Reference(
            ws, min_col=5 + len(duty_cycle_list) + 6, max_col=4 + len(duty_cycle_list) + 6 + len(duty_cycle_list), min_row=4, max_row=4 + len(frequency_list))
        labels = Reference(
            ws, min_col=4 + len(duty_cycle_list) + 6, min_row=5, max_row=4 + len(frequency_list))
        vmax_surface.add_data(ref, titles_from_data=True)
        vmax_surface.set_categories(labels)
        vmax_surface.title = "3D VMAX"
        ws.add_chart(vmax_surface, "{}{}".format(
            vmax_graph_col, 5 + len(frequency_list) + 6))

    wb.save(filename=output_file_name)
    # Pause with no load for the specified time before

    # Switch back to SVDC mode and turn off the load...
    generator_api.Generator1SVSC(target_rail_1, 0, False)
    generator_api.Generator2SVSC(target_rail_2, 0, False)
    generator_api.Generator3SVSC(target_rail_3, 0, False)
    generator_api.Generator4SVSC(target_rail_4, 0, False)

    data_api.SetFanSpeed(1)
    measurement_api.ClearAllMeasurements()
    display_api.Ch1_2Rail("NOLOAD")
    display_api.SetChannel(ScoplessChannel.Ch1, False)
    display_api.SetChannel(ScoplessChannel.Ch2, False)

    display_api.Ch3_4Rail("NOLOAD")
    display_api.SetChannel(ScoplessChannel.Ch3, False)
    display_api.SetChannel(ScoplessChannel.Ch4, False)

    print("Test completed. Data saved to:  ", output_file_name)


if __name__ == "__main__":
    fivra_3d(target_voltage=1.8,
             start_frequency=1,
             end_frequency=1000,
             sample_per_decade=5,
             nw_current_low=4.7,
             nw_current_high=13.4,
             ne_current_low=2.85,
             ne_current_high=7.53,
             sw_current_low=2.97,
             sw_current_high=8.85,
             se_current_low=7,
             se_current_high=18.3,
             nw_slew_rate=10.8,
             ne_slew_rate=12.1,
             sw_slew_rate=22,
             se_slew_rate=10.9,
             cooldown_delay=0,
             duty_cycle_list=[10, 20, 30, 40, 50, 60, 70, 80, 90],
             debug=True,
             )
