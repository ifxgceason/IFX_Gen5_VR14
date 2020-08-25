# eason.chen@infineon.com 2020/Aug/10
import clr
import sys
import os
import ifx
dll_path=r"C:\GEN5\PythonScripts"
sys.path.append(dll_path)

import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import pandas as pd
import datetime

from Common.APIs import DataAPI
from Common.APIs import DisplayAPI
from Common.Models import TriggerModel
from Common.Models import FanModel
from Common.Models import RawDataModel
from Common.Models import DataModel
from Common.Models import DisplayModel
from Common.Enumerations import HorizontalScale
from Common.Enumerations import ScoplessChannel
from Common.Enumerations import Cursors
from Common.APIs import MeasurementAPI
from Common.APIs import GeneratorAPI
# from Common.APIs import MeasurementItemsAPI
#from Common.Models import MeasurementItemModel
#from CommonBL import *
from Common.Enumerations import PowerState
from Common.Enumerations import Transition
from Common.Enumerations import PowerState
from Common.Enumerations import Transition
from Common.Enumerations import Protocol
from Common.Enumerations import RailTestMode
# from Common.Enumerations import MeasurementItemType
# from Common.Enumerations import MeasurementItemUnit

from System import String, Char, Int32, UInt16, Boolean, Array, Byte, Double
from System.Collections.Generic import List
from System.Collections.Generic import Dictionary

gen = GeneratorAPI();
rails = gen.GetRails();
dat=DataAPI()
data=DataAPI()
display = DisplayAPI()
Measurment=MeasurementAPI()
#items = MeasurementItemsAPI()

from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

def vr14_3d(rail_name="VCCIN",
            vout_list=[1.83,1.73],
            start_freq=1,
            end_freq=1000,
            icc_min=1,
            icc_max=100,
            freq_steps_per_decade=5,
            rise_time=262,
            cool_down_delay=0,
            start_duty=10,
            end_duty=50,
            duty_step=10,
            excel=True):
    for vout in vout_list:
        MeasurementAPI().ClearAllMeasurements()    
        print(f"dll verison = {ifx.version()}")
        #detect Infinon device
        svid_addr,svid_bus=ifx.rail_name_to_svid_parameter(rail_name)
        ifx.getsvidregvalue(svid_addr,svid_bus,0x15)
        duty_cycle_list=list(range(start_duty,end_duty+1,duty_step))
        ScriptVersion = 1.1

        # Set up frequency step list (all frequencies are in kHz):
        FrequencyList=[]
        StartOfDecade = float(start_freq)
        CurrentFreq = StartOfDecade
        print("Frequency set points:")
        while CurrentFreq < float(end_freq):
            if (StartOfDecade * 10.0) > float(end_freq):
                EndOfDecade = float(end_freq)
            else:
                EndOfDecade = float(StartOfDecade) * 10.0
                print("New decade from ", StartOfDecade, "kHz to ", EndOfDecade, "kHz")
                
            StepSize = (float(EndOfDecade) - float(StartOfDecade)) / (freq_steps_per_decade - 1)
            print("Step size = ", StepSize, "kHz")
            
            for j in range(freq_steps_per_decade - 1):
                FrequencyList.append(CurrentFreq)
                print("    ", CurrentFreq, "kHz")
                CurrentFreq = CurrentFreq + StepSize
                
            # Move on to the next frequency decade...
            StartOfDecade = CurrentFreq

        # Add the final frequency setpoint...
        if CurrentFreq > float(end_freq):
            CurrentFreq = float(end_freq)
        FrequencyList.append(CurrentFreq)
        print("    ", CurrentFreq, "kHz")
        print("There will be ", len(duty_cycle_list), " Duty Cycle data points per frequency setpoint.")
        print("Total of ", len(FrequencyList) * len(duty_cycle_list), " data points will be taken.")
        #////////////////////////////////////////////////////////////////////////////

        # Set up the first load tab to work with out target rail and display it on the scope view...
        gen.AssignRailToDriveOne(rail_name)    
        display.Ch1_2Rail(rail_name)
        display.SetChannel(ScoplessChannel.Ch1, True)
        display.SetChannel(ScoplessChannel.Ch2, True)
        #Set vertical scale (the target voltage +/- 25%)...
        display.SetVerticalVoltageScale(vout * 0.75, vout * 1.25)
        display.SetVerticalCurrentScale(0, (icc_max * 1.1))

        #
        # SetTrigger(<source>,<display_pos>,<DAC_trig_pos>,<ADC_trig_level>,<ADC_trig_slope>,<RESERVED>)
        #
        # Set up our trigger for DAC1 (Load tab #1 drive signal), negative slope, with trigger on the left of the scope display...
        dat.SetTrigger(4, 0, 0, 0, 0, False)

        # Set the rail voltage...
        gen.SetVoltageForRail(rail_name, vout, Transition.Fast)   
        time.sleep(1)


        # Set up all required measurements...
        Measurment.MeasureCurrentFrequency(rail_name)
        Measurment.MeasureCurrentDutyCycle(rail_name)
        Measurment.MeasureVoltageAmplitude(rail_name)
        Measurment.MeasureVoltageOverUnderShoot(rail_name)
        Measurment.MeasureVoltageMinMax(rail_name)


        #Create the output Excel file...
        #wb=Workbook()
        #ws1 = wb.create_sheet(title="3D", index=1)

        theTime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name="IFX-VR-"+rail_name+"-3D-"+str(vout)+"V-"+str(start_freq)+"Khz-"+str(end_freq)+"khz-"+theTime+".xlsx"

        # Set Fan Speed to Max...
        dat.SetFanSpeed(99)

        result_temp=dict()
        result=pd.DataFrame()
        for i in range (0, len(FrequencyList)):
            for dc_index in range (0, len(duty_cycle_list)):
                # Set the horizontal scale such that at least 2 complete cycles are displayed at the current frequency...
                # Get minimum period x5 (convert to microseconds)...
                MinimumTime = (1.0 / float(FrequencyList[i])) * 500
                if MinimumTime >= 5000:
                    display.SetHorizontalScale(HorizontalScale.Scale5ms)
                    print("Scale = 5ms (MinTime =", MinimumTime, ")")
                elif MinimumTime >= 2000:
                    display.SetHorizontalScale(HorizontalScale.Scale2ms)
                    print("Scale = 2ms (MinTime =", MinimumTime, ")")
                elif MinimumTime >= 1000:
                    display.SetHorizontalScale(HorizontalScale.Scale1ms)
                    print("Scale = 1ms (MinTime =", MinimumTime, ")")
                elif MinimumTime >= 400:
                    display.SetHorizontalScale(HorizontalScale.Scale500us)
                    print("Scale = 500ns (MinTime =", MinimumTime,")")
                elif MinimumTime >= 200:
                    display.SetHorizontalScale(HorizontalScale.Scale200us)
                    print("Scale = 200ns (MinTime =", MinimumTime,")")
                elif MinimumTime >= 80:
                    display.SetHorizontalScale(HorizontalScale.Scale80us)
                    print("Scale = 80ns (MinTime =", MinimumTime,")")
                elif MinimumTime >= 40:
                    display.SetHorizontalScale(HorizontalScale.Scale40us)
                    print("Scale = 40ns (MinTime =", MinimumTime,")")
                elif MinimumTime >= 20:
                    display.SetHorizontalScale(HorizontalScale.Scale20us)
                    print("Scale = 20ns (MinTime =", MinimumTime,")")
                else:
                    display.SetHorizontalScale(HorizontalScale.Scale10us)
                    print("Scale = 10ns (MinTime =", MinimumTime,")")
                
                # Set the load current step size, frequency, ramp time, and duty cycle...
                gen.Generator1SVDC(rail_name, icc_max, icc_min, FrequencyList[i], duty_cycle_list[dc_index], rise_time, True)

                # Wait 1 second for measurements to catch up...
                time.sleep(1)

                # Measure the actual frequency. This will account for any rounding done by the Gen5 API...
                FrequencyString = Measurment.GetCurrentFrequencyOnce(rail_name)
                #Frequency = items.GetCurrentFrequencyOnce(rail_name)[0].Value
                Frequency = FrequencyString[:FrequencyString.find('k')]

                # Measure the amplitude...
                VppString = Measurment.GetVoltageAmplitudeOnce(rail_name)
                #Vpp = items.GetVoltageAmplitudeOnce(rail_name)[0].Value
                Vpp = VppString[:VppString.find('V')]

                # Measure overshoot and undershoot...
                VOverUnderString = Measurment.GetVoltageOverUnderShootsOnce(rail_name)
                #VOvershoot = items.GetVoltageOverUnderShootsOnce(rail_name)[0].Value
                VOvershoot = VOverUnderString[VOverUnderString.find("Over")+7:VOverUnderString.find('V')]
                #VUndershoot = items.GetVoltageOverUnderShootsOnce(rail_name)[1].Value
                VUndershoot = VOverUnderString[VOverUnderString.find("Under")+8:VOverUnderString.find('V', VOverUnderString.find("Under"),len(VOverUnderString))]

                # Measure VMin and VMax...
                VMinMax = Measurment.GetVoltageMinMaxOnce(rail_name)
                VMax = VMinMax[7:VMinMax.find('V')]
                VMin = VMinMax[VMinMax.find('Low') + 6:VMinMax.find('V', VMinMax.find('Low'))]
                
                # Show each data point taken in script output window...
                #print("Frequency: ", FrequencyList[i], "   DutyCycle: ", duty_cycle_list[dc_index], "%  VOvershoot: " , VOvershoot, "V   VUndershoot: ", VUndershoot, "V   Vp-p: ", Vpp,"V")

                result_temp['Frequency']=float(FrequencyList[i])
                result_temp['DutyCycleList']=float(duty_cycle_list[dc_index])
                result_temp['VOvershoot']=float(VOvershoot)
                result_temp['VUndershoot']=float(VUndershoot)
                result_temp['Vpp']=float(Vpp)
                result_temp['VMin']=float(VMin)
                result_temp['VMax']=float(VMax)

                result=result.append(result_temp,ignore_index=True)        
                print(result)          
            # Pause with no load for the specified time before advancing to the next frequency step...
            print("Cool-down delay...")
            gen.Generator1SVDC(rail_name, icc_max, icc_min, FrequencyList[i], duty_cycle_list[dc_index], rise_time, False)
            time.sleep(cool_down_delay)
        # Next Frequency

        df=pd.DataFrame(result,columns=["Frequency","DutyCycleList","VOvershoot","VUndershoot","Vpp","VMin","VMax"])
        df_vmax=df.sort_values(['VMax'],ascending=False)
        df_vmin=df.sort_values(['VMin'],ascending=True)
        # Switch back to SVDC mode and turn off the load...
        gen.Generator1SVSC(rail_name, 0, False)
        dat.SetFanSpeed(1)

        # Save the output file...
        ifx.df1_df2_to_excel(df_vmax,df_vmin,file_name,sheet_name1="Vmax",sheet_name2="Vmin")
        #df.to_excel(file_name)

        print("Test completed. Data saved to:  ", file_name)
        ifx.excelSelect(file_name,excel)
        ifx.ifxPlt(df)
    print("3D completed")

    
if __name__ == "__main__":
    vr14_3d(rail_name="VCCIN",
            vout_list=[1.83],
            start_freq=0.1,
            end_freq=20,
            icc_min=61,
            icc_max=405,
            freq_steps_per_decade=7,
            rise_time=2000,
            cool_down_delay=0,
            start_duty=10,
            end_duty=50,
            duty_step=10,
            excel=True)
